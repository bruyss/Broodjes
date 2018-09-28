#! python3
# broodjes.py

import smtplib
import imapclient
import datetime
import time
import pyzmail
import openpyxl
from openpyxl.styles import Font
import os.path as op
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from email import encoders
import logging

logging.basicConfig(filename='log.txt',
                    level=logging.WARNING,
                    format=' %(asctime)s - %(levelname)s- %(message)s')

address = 'bachelorproefgroep063@gmail.com'
pw = 'R4t&756L'
verantwoordelijke = 'jeroen.vb1@gmail.com'
today = datetime.date.today()
deadline = datetime.time(9, 30)
subjects = ('Broodje', 'broodje', 'Broodjes', 'broodjes', 'Bestelling broodje', 'Bestelling broodjes', 'bestelling broodje', 'bestelling broodjes')


def sendmail(send_from, send_to, subject, message, files=[],
              server="localhost", port=587, username='', password='',
              use_tls=True):
    """Compose and send email with provided info and attachments.

    Args:
        send_from (str): from name
        send_to (str): to name
        subject (str): message title
        message (str): message body
        files (list[str]): list of file paths to be attached to email
        server (str): mail server host name
        port (int): port number
        username (str): server auth username
        password (str): server auth password
        use_tls (bool): use TLS mode
    """
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = send_to
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(message))

    for path in files:
        part = MIMEBase('application', "octet-stream")
        with open(path, 'rb') as file:
            part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        'attachment; filename="{}"'.format(op.basename(path)))
        msg.attach(part)

    smtp = smtplib.SMTP(server, port)
    if use_tls:
        smtp.starttls()
    smtp.login(username, password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()


def getmailIDs(subject):        # Geeft mail IDs van ongelezen emails die vandaag ontvangen zijn met bepaald onderwerp
    imapObj = imapclient.IMAPClient('imap.gmail.com', ssl=True)
    imapObj.login(address, pw)
    imapObj.select_folder('INBOX', readonly=False)
    UIds = imapObj.search(['SINCE', today, 'SUBJECT', subject, 'UNSEEN'])
    return imapObj, UIds


def getmail(imapObj, ID):       # Geeft verstuurder en eerste lijn van body van een email
    rMessage = imapObj.fetch([ID], ['BODY[]', 'FLAGS'])
    message = pyzmail.PyzMessage.factory(rMessage[ID][b'BODY[]'])
    address = message.get_addresses('from')[0]
    body = message.text_part.get_payload().decode(message.text_part.charset)
    firstline = body.split('\r\n')[0]
    return address, firstline


def initworksheet():        # Initialisering van een nieuwe excel voor bestellingen
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Bestellingen'
    headingfont = Font(bold=True)
    sheet['A1'].font = headingfont
    sheet.column_dimensions['A'].width = 30
    sheet['A1'] = 'Naam'
    sheet['B1'].font = headingfont
    sheet.column_dimensions['B'].width = 30
    sheet['B1'] = 'Mailadres'
    sheet['C1'].font = headingfont
    sheet.column_dimensions['C'].width = 40
    sheet['C1'] = 'Bestelling'
    wsnaam = 'Broodjes%s.xlsx' % today.isoformat()
    wb.save('./data/' + wsnaam)
    logging.warning('Nieuwe excel %s aangemaakt.' % wsnaam)
    return wsnaam


def getnieuwebestellingen():        # Haalt alle nieuwe bestellingen van de mailserver
    nieuwebestellers = []
    nieuwebestellingen = []
    imapObj, IDs = getmailIDs('Broodje')
    for ID in IDs:
        besteller, bestelling = getmail(imapObj, ID)
        nieuwebestellers.append(besteller)
        nieuwebestellingen.append(bestelling)
        logging.warning('Bestelling ontvangen van %s: %s voor %s' % (besteller[1], bestelling, besteller[0]))
        print('%s voor %s vanuit %s' % (bestelling, besteller[0], besteller[1]))
        sendmail('Robobroodje', besteller[1],
                 'Bestelling broodje',
                 'Beste %s\n\nU bestelling van %s is goed ontvangen en zal deze middag klaarliggen.\n\nMvg\nRobobroodje' % (besteller[0], bestelling),
                 [],
                 'smtp.gmail.com',
                 587,
                 address,
                 pw)
    imapObj.logout()
    return nieuwebestellers, nieuwebestellingen


def vulexcel(wsnaam, bestellers, bestellingen):     # Vult excel aan met bestellingen
    wb = openpyxl.load_workbook('./data/' + wsnaam)
    sheet = wb.active
    start = sheet.max_row
    for i in range(len(bestellers)):
        rnum = start + i + 1
        sheet['A' + str(rnum)] = bestellers[i][0]
        sheet['B' + str(rnum)] = bestellers[i][1]
        sheet['C' + str(rnum)] = bestellingen[i]
    wb.save('./data/' + wsnaam)


def stuurbestelling(wsnaam, ontvanger):
    sendmail('Robobroodje',
             ontvanger,
             'Bestellingen',
             'In bijlage de besetelling voor vandaag\n\nMvg\nRobobroodje',
             ['./data/' + wsnaam, 'log.txt'],
             'smtp.gmail.com',
             587,
             address,
             pw)


def sleeper():
    morgen = datetime.datetime.now() + datetime.timedelta(days=1)
    morgenvroeg = datetime.datetime(morgen.year, morgen.month, morgen.day + 1, 5)
    while datetime.datetime.now() < morgenvroeg:
        time.sleep(10)


wsnaam = initworksheet()
while True:
    nieuwebestellers, nieuwebestellingen = getnieuwebestellingen()
    vulexcel(wsnaam, nieuwebestellers, nieuwebestellingen)
    if int(datetime.datetime.now().strftime('%H%M')) > 1330:
        logging.warning('Bestelling gestuurd, tot morgen :)')
        stuurbestelling(wsnaam, verantwoordelijke)
        sleeper()
        wsnaam = initworksheet()
    time.sleep(10)
